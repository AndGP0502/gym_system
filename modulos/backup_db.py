import os
import sqlite3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from tkinter import filedialog, messagebox
from modulos.rutas import get_db_path, get_config_path, get_backup_dir

DB_PATH        = get_db_path()
CONFIG_PATH    = get_config_path()
CARPETA_BACKUP = get_backup_dir()

TABLAS = {
    "clientes": [
        "id", "nombre", "cedula", "telefono",
        "fecha_registro", "edad", "fecha_nacimiento", "sexo"
    ],
    "membresias": [
        "id", "nombre_plan", "precio", "duracion_dias"
    ],
    "suscripciones": [
        "id", "cliente_id", "membresia_id", "fecha_inicio",
        "fecha_vencimiento", "precio_total", "pagado", "pendiente"
    ],
    "pagos": [
        "id", "suscripcion_id", "monto", "fecha_pago"
    ],
    "ficha_cliente": [
        "id", "cliente_id", "objetivo", "estado_fisico", "condiciones",
        "notas", "foto_ruta", "peso_kg", "altura_m", "cir_abdominal",
        "status_fisico", "objetivo_2", "peso_ideal", "lesion",
        "cardiovascular", "asfixia", "asmatico", "medicacion", "mareos"
    ],
    "historial_medidas": [
        "id", "cliente_id", "fecha", "peso_kg", "altura_cm", "imc", "notas"
    ],
    "alertas_enviadas": [
        "id_cliente", "fecha_alerta"
    ],
}

COLORES = {
    "clientes":          "4F81BD",
    "membresias":        "9BBB59",
    "suscripciones":     "F79646",
    "pagos":             "8064A2",
    "ficha_cliente":     "4BACC6",
    "historial_medidas": "C0504D",
    "alertas_enviadas":  "7F7F7F",
}


def _estilo_encabezado(ws, columnas: list, color_hex: str):
    fill   = PatternFill("solid", fgColor=color_hex)
    fuente = Font(bold=True, color="FFFFFF")
    for col_idx, nombre_col in enumerate(columnas, start=1):
        celda = ws.cell(row=1, column=col_idx, value=nombre_col)
        celda.fill      = fill
        celda.font      = fuente
        celda.alignment = Alignment(horizontal="center")
        ws.column_dimensions[celda.column_letter].width = max(len(nombre_col) + 4, 14)


def exportar_backup_excel() -> str | None:
    os.makedirs(CARPETA_BACKUP, exist_ok=True)

    ruta_destino = filedialog.asksaveasfilename(
        title="Guardar backup como Excel",
        defaultextension=".xlsx",
        filetypes=[("Excel", "*.xlsx")],
        initialdir=CARPETA_BACKUP,
        initialfile=f"gym_backup_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
    )
    if not ruta_destino:
        return None

    try:
        con = sqlite3.connect(DB_PATH)
        wb  = openpyxl.Workbook()
        wb.remove(wb.active)

        for tabla, columnas in TABLAS.items():
            ws = wb.create_sheet(title=tabla)
            _estilo_encabezado(ws, columnas, COLORES.get(tabla, "4F81BD"))
            try:
                filas = con.execute(
                    f"SELECT {', '.join(columnas)} FROM {tabla}"
                ).fetchall()
                for fila in filas:
                    ws.append(list(fila))
            except Exception as e:
                ws.cell(row=2, column=1, value=f"[ERROR al leer tabla: {e}]")

        con.close()
        wb.save(ruta_destino)
        messagebox.showinfo(
            "Backup Excel creado",
            f"Backup guardado exitosamente en:\n{ruta_destino}"
        )
        return ruta_destino

    except Exception as e:
        import traceback
        messagebox.showerror(
            "Error al exportar",
            f"No se pudo crear el backup Excel:\n{e}\n\n{traceback.format_exc()}"
        )
        return None


def importar_desde_excel() -> bool:
    archivo = filedialog.askopenfilename(
        title="Seleccionar backup Excel a importar",
        filetypes=[("Excel", "*.xlsx")],
        initialdir=CARPETA_BACKUP if os.path.exists(CARPETA_BACKUP) else "/"
    )
    if not archivo:
        return False

    try:
        wb  = openpyxl.load_workbook(archivo, read_only=True, data_only=True)
        con = sqlite3.connect(DB_PATH)
        cur = con.cursor()
        errores = []

        for tabla, columnas in TABLAS.items():
            if tabla not in wb.sheetnames:
                errores.append(f"Hoja '{tabla}' no encontrada en el Excel.")
                continue

            ws   = wb[tabla]
            rows = list(ws.iter_rows(values_only=True))

            if len(rows) < 2:
                continue

            encabezados_excel = [str(c).strip() if c else "" for c in rows[0]]
            if encabezados_excel != columnas:
                errores.append(
                    f"Columnas de '{tabla}' no coinciden.\n"
                    f"  Esperado:  {columnas}\n"
                    f"  En Excel:  {encabezados_excel}"
                )
                continue

            placeholders = ", ".join(["?"] * len(columnas))
            cols_str     = ", ".join(columnas)
            sql          = f"INSERT OR REPLACE INTO {tabla} ({cols_str}) VALUES ({placeholders})"

            for fila in rows[1:]:
                if all(v is None for v in fila):
                    continue
                if str(fila[0]).startswith("[ERROR"):
                    continue
                try:
                    cur.execute(sql, list(fila))
                except Exception as e:
                    errores.append(f"Fila en '{tabla}': {e} | datos: {fila}")

        con.commit()
        con.close()
        wb.close()

        if errores:
            messagebox.showwarning(
                "Importación con advertencias",
                f"Se importaron datos con los siguientes problemas:\n\n" +
                "\n".join(errores[:10])
            )
        else:
            messagebox.showinfo(
                "Importación exitosa",
                "Todos los datos fueron restaurados correctamente desde el Excel."
            )
        return True

    except Exception as e:
        import traceback
        messagebox.showerror(
            "Error al importar",
            f"No se pudo importar el backup:\n{e}\n\n{traceback.format_exc()}"
        )
        return False


def crear_backup() -> str:
    os.makedirs(CARPETA_BACKUP, exist_ok=True)
    fecha       = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    ruta_backup = os.path.join(CARPETA_BACKUP, f"gym_backup_{fecha}.db")

    con_origen = sqlite3.connect(DB_PATH)
    con_backup = sqlite3.connect(ruta_backup)
    con_origen.backup(con_backup, pages=-1)
    con_backup.close()
    con_origen.close()

    return ruta_backup