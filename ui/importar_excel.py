import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from tkinter import filedialog, messagebox
import openpyxl
import sqlite3
import os
from datetime import datetime

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH  = os.path.join(BASE_DIR, "..", "gym.db")

# ── Planes FIVGYM ─────────────────────────────────────────────────────────────
PLANES_FIVGYM = {
    "MENSUAL /30 DIAS":              {"duracion": 30,  "precio": 30},
    "TRIMESTRAL / 90 DIAS":          {"duracion": 90,  "precio": 80},
    "SEMESTRAL / 180 DIAS":          {"duracion": 180, "precio": 150},
    "SEMESTRAL / 188 DIAS":          {"duracion": 188, "precio": 150},
    "ANUAL / 365 DIAS":              {"duracion": 365, "precio": 250},
    "PLAN PAREJA / X2 / 30 DIAS":    {"duracion": 30,  "precio": 25},
    "PLAN ESTUDIANTIL / 30 DIAS":    {"duracion": 30,  "precio": 25},
    "PLAN 3RA EDAD / 30 DIAS / 50%": {"duracion": 30,  "precio": 20},
}


def _con():
    return sqlite3.connect(DB_PATH)


def _init_tablas_extra():
    """Asegura que existan las columnas y tablas extra que necesita la importación."""
    con = _con()
    # Columna edad en clientes
    for col, tipo in [("edad", "REAL"), ("fecha_nacimiento", "TEXT"), ("sexo", "TEXT")]:
        try:
            con.execute(f"ALTER TABLE clientes ADD COLUMN {col} {tipo}")
        except sqlite3.OperationalError:
            pass
    # Tabla ficha_cliente con condiciones médicas
    con.execute("""
        CREATE TABLE IF NOT EXISTS ficha_cliente (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            cliente_id      INTEGER UNIQUE,
            objetivo        TEXT,
            estado_fisico   TEXT,
            condiciones     TEXT,
            notas           TEXT,
            foto_ruta       TEXT,
            FOREIGN KEY(cliente_id) REFERENCES clientes(id)
        )
    """)
    # Columnas extra en ficha_cliente
    for col, tipo in [
        ("peso_kg",        "REAL"),
        ("altura_m",       "REAL"),
        ("cir_abdominal",  "REAL"),
        ("status_fisico",  "TEXT"),
        ("objetivo_2",     "TEXT"),
        ("peso_ideal",     "REAL"),
        ("lesion",         "TEXT"),
        ("cardiovascular", "TEXT"),
        ("asfixia",        "TEXT"),
        ("asmatico",       "TEXT"),
        ("medicacion",     "TEXT"),
        ("mareos",         "TEXT"),
    ]:
        try:
            con.execute(f"ALTER TABLE ficha_cliente ADD COLUMN {col} {tipo}")
        except sqlite3.OperationalError:
            pass
    con.commit()
    con.close()

_init_tablas_extra()


def _parsear_fecha(valor) -> str | None:
    if valor is None:
        return None
    if isinstance(valor, datetime):
        if valor.year < 1990:
            return None
        return valor.strftime("%Y-%m-%d")
    txt = str(valor).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(txt, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def _parsear_numero(valor) -> float | None:
    if valor is None:
        return None
    try:
        return float(str(valor).replace(",", ".").strip())
    except Exception:
        return None


def _obtener_o_crear_membresia(nombre_plan: str, precio: float, duracion: int) -> int:
    con = _con()
    cur = con.cursor()

    # Detectar qué columna usa esta BD para el nombre del plan
    cur.execute("PRAGMA table_info(membresias)")
    cols = [row[1] for row in cur.fetchall()]
    col_nombre = "nombre_plan" if "nombre_plan" in cols else "nombre"

    cur.execute(f"SELECT id FROM membresias WHERE {col_nombre} = ?", (nombre_plan,))
    fila = cur.fetchone()
    if fila:
        con.close()
        return fila[0]

    cur.execute(
        f"INSERT INTO membresias ({col_nombre}, precio, duracion_dias) VALUES (?, ?, ?)",
        (nombre_plan, precio, duracion)
    )
    id_mem = cur.lastrowid
    con.commit()
    con.close()
    return id_mem


# ── IMPORTAR PLANTILLA ESTÁNDAR ───────────────────────────────────────────────

def importar_plantilla(ruta: str) -> dict:
    wb = openpyxl.load_workbook(ruta, data_only=True)
    resultados = {"clientes": 0, "membresias": 0, "errores": []}

    if "Membresias" in wb.sheetnames:
        ws = wb["Membresias"]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < 2: continue
            nombre, precio, duracion = row[0], row[1], row[2]
            if not nombre: continue
            try:
                _obtener_o_crear_membresia(str(nombre).strip(), float(precio or 0), int(duracion or 30))
                resultados["membresias"] += 1
            except Exception as e:
                resultados["errores"].append(f"Membresía '{nombre}': {e}")

    if "Clientes" in wb.sheetnames:
        ws = wb["Clientes"]
        con = _con()
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < 2: continue
            nombre, cedula, telefono, fecha = row[0], row[1], row[2], row[3]
            if not nombre: continue
            fecha_str = _parsear_fecha(fecha) or datetime.now().strftime("%d/%m/%Y")
            try:
                con.execute(
                    "INSERT OR IGNORE INTO clientes (nombre, cedula, telefono, fecha_registro) VALUES (?,?,?,?)",
                    (str(nombre).strip(), str(cedula or "").strip(), str(telefono or "").strip(), fecha_str)
                )
                resultados["clientes"] += 1
            except Exception as e:
                resultados["errores"].append(f"Cliente '{nombre}': {e}")
        con.commit()
        con.close()

    return resultados


# ── IMPORTAR TABLA FIVGYM ─────────────────────────────────────────────────────

def importar_fivgym(ruta: str) -> dict:
    wb = openpyxl.load_workbook(ruta, data_only=True)
    resultados = {"clientes": 0, "fichas": 0, "suscripciones": 0, "errores": []}

    # ── Una sola conexión para todo el proceso ────────────────────────────────
    con = _con()
    con.execute("PRAGMA journal_mode=WAL")
    con.execute("PRAGMA foreign_keys = OFF")

    # Detectar columna nombre de membresías
    cols_mem = [r[1] for r in con.execute("PRAGMA table_info(membresias)").fetchall()]
    col_nombre_mem = "nombre_plan" if "nombre_plan" in cols_mem else "nombre"

    def obtener_o_crear_mem_local(nombre_plan, precio, duracion):
        fila = con.execute(
            f"SELECT id FROM membresias WHERE {col_nombre_mem} = ?", (nombre_plan,)
        ).fetchone()
        if fila:
            return fila[0]
        cur = con.execute(
            f"INSERT INTO membresias ({col_nombre_mem}, precio, duracion_dias) VALUES (?,?,?)",
            (nombre_plan, precio, duracion)
        )
        return cur.lastrowid

    # ── PASO 1: leer DATOS INFORMATIVOS ──────────────────────────────────────
    ws_info = wb["DATOS INFORMATIVOS"]
    info_por_num = {}
    for i, row in enumerate(ws_info.iter_rows(values_only=True)):
        if i < 3: continue
        num = row[0]
        if not num or not isinstance(num, int): break
        nombre = str(row[3] or "").strip()
        if not nombre or nombre == "0": continue
        fecha_nac = _parsear_fecha(row[4])
        edad_raw  = row[5]
        edad = round(float(edad_raw), 1) if edad_raw and isinstance(edad_raw, (int, float)) and float(edad_raw) < 120 else None
        info_por_num[num] = {
            "nombre":         nombre,
            "telefono":       str(int(row[1])).strip() if row[1] else "",
            "cedula":         str(int(row[2])).strip() if row[2] else "",
            "fecha_nac":      fecha_nac,
            "edad":           edad,
            "sexo":           str(row[6] or "").strip(),
            "lesion":         str(row[7]  or "NO").strip(),
            "cardiovascular": str(row[8]  or "NO").strip(),
            "asfixia":        str(row[9]  or "NO").strip(),
            "asmatico":       str(row[10] or "NO").strip(),
            "medicacion":     str(row[11] or "NO").strip(),
            "mareos":         str(row[12] or "NO").strip(),
        }

    # ── PASO 2: leer DATOS FISICOS ────────────────────────────────────────────
    ws_fis = wb["DATOS FISICOS"]
    fisicos_por_num = {}
    for i, row in enumerate(ws_fis.iter_rows(values_only=True)):
        if i < 3: continue
        num = row[0]
        if not num or not isinstance(num, int): break
        fisicos_por_num[num] = {
            "peso_kg":       _parsear_numero(row[3]),
            "altura_m":      _parsear_numero(row[4]),
            "cir_abdominal": _parsear_numero(row[5]),
            "status":        str(row[6] or "").strip(),
            "objetivo1":     str(row[7] or "").strip(),
            "objetivo2":     str(row[8] or "").strip(),
            "peso_ideal":    _parsear_numero(row[9]),
        }

    # ── PASO 3: leer fechas reales de PLANES Y PAGOS ──────────────────────────
    ws_pagos = wb["PLANES Y PAGOS"]
    fecha_real_por_num = {}
    suscripciones_excel = []
    for i, row in enumerate(ws_pagos.iter_rows(values_only=True)):
        if i < 3: continue
        num = row[0]
        if not num or not isinstance(num, int): break
        nombre = str(row[1] or "").strip()
        plan   = str(row[2] or "").strip()
        fecha_inicio = _parsear_fecha(row[6])
        fecha_fin    = _parsear_fecha(row[7])
        if fecha_inicio and num not in fecha_real_por_num:
            fecha_real_por_num[num] = fecha_inicio
        if plan and nombre and fecha_inicio:
            suscripciones_excel.append({
                "num":    num,
                "nombre": nombre,
                "plan":   plan,
                "pago1":  float(row[3] or 0),
                "pago2":  float(row[4] or 0),
                "total":  float(row[5] or 0),
                "inicio": fecha_inicio,
                "fin":    fecha_fin,
            })

    # ── PASO 4: insertar clientes ─────────────────────────────────────────────
    clientes_map = {}
    for num, info in info_por_num.items():
        fecha_inicio_real = fecha_real_por_num.get(num)
        if fecha_inicio_real:
            try:
                dt = datetime.strptime(fecha_inicio_real, "%Y-%m-%d")
                fecha_reg = dt.strftime("%d/%m/%Y")
            except Exception:
                fecha_reg = fecha_inicio_real
        else:
            fecha_reg = datetime.now().strftime("%d/%m/%Y")

        try:
            con.execute("""
                INSERT OR IGNORE INTO clientes
                    (nombre, cedula, telefono, fecha_registro, edad, fecha_nacimiento, sexo)
                VALUES (?,?,?,?,?,?,?)
            """, (info["nombre"], info["cedula"], info["telefono"],
                  fecha_reg, info["edad"], info["fecha_nac"], info["sexo"]))

            fila = con.execute(
                "SELECT id FROM clientes WHERE nombre = ? ORDER BY id DESC LIMIT 1",
                (info["nombre"],)
            ).fetchone()
            if fila:
                clientes_map[num] = fila[0]
                resultados["clientes"] += 1
        except Exception as e:
            resultados["errores"].append(f"Cliente '{info['nombre']}': {e}")

    # ── PASO 5: insertar fichas ───────────────────────────────────────────────
    for num, id_cliente in clientes_map.items():
        info = info_por_num.get(num, {})
        fis  = fisicos_por_num.get(num, {})
        condiciones_list = []
        if info.get("lesion",         "").upper() == "SI": condiciones_list.append("Lesion muscular/articular")
        if info.get("cardiovascular", "").upper() == "SI": condiciones_list.append("Enfermedad cardiovascular")
        if info.get("asfixia",        "").upper() == "SI": condiciones_list.append("Se asfixia con facilidad")
        if info.get("asmatico",       "").upper() == "SI": condiciones_list.append("Asmatico/Epileptico/Diabetico")
        if info.get("medicacion",     "").upper() == "SI": condiciones_list.append("Toma medicacion")
        if info.get("mareos",         "").upper() == "SI": condiciones_list.append("Mareos/desmayos al ejercitar")
        condiciones_str = ", ".join(condiciones_list) if condiciones_list else "Ninguna"
        try:
            con.execute("""
                INSERT INTO ficha_cliente
                    (cliente_id, objetivo, estado_fisico, condiciones, notas, foto_ruta,
                     peso_kg, altura_m, cir_abdominal, status_fisico, objetivo_2, peso_ideal,
                     lesion, cardiovascular, asfixia, asmatico, medicacion, mareos)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                ON CONFLICT(cliente_id) DO UPDATE SET
                    objetivo=excluded.objetivo, estado_fisico=excluded.estado_fisico,
                    condiciones=excluded.condiciones, peso_kg=excluded.peso_kg,
                    altura_m=excluded.altura_m, cir_abdominal=excluded.cir_abdominal,
                    status_fisico=excluded.status_fisico, objetivo_2=excluded.objetivo_2,
                    peso_ideal=excluded.peso_ideal, lesion=excluded.lesion,
                    cardiovascular=excluded.cardiovascular, asfixia=excluded.asfixia,
                    asmatico=excluded.asmatico, medicacion=excluded.medicacion,
                    mareos=excluded.mareos
            """, (
                id_cliente,
                fis.get("objetivo1", ""), fis.get("status", ""),
                condiciones_str, "", None,
                fis.get("peso_kg"), fis.get("altura_m"), fis.get("cir_abdominal"),
                fis.get("status", ""), fis.get("objetivo2", ""), fis.get("peso_ideal"),
                info.get("lesion","NO"), info.get("cardiovascular","NO"),
                info.get("asfixia","NO"), info.get("asmatico","NO"),
                info.get("medicacion","NO"), info.get("mareos","NO"),
            ))
            resultados["fichas"] += 1
        except Exception as e:
            resultados["errores"].append(f"Ficha {id_cliente}: {e}")

    # ── PASO 6: crear membresías e insertar suscripciones ────────────────────
    for sus in suscripciones_excel:
        num    = sus["num"]
        nombre = sus["nombre"]
        plan   = sus["plan"]

        id_cliente = clientes_map.get(num)
        if not id_cliente:
            fila = con.execute(
                "SELECT id FROM clientes WHERE nombre = ? ORDER BY id DESC LIMIT 1",
                (nombre,)
            ).fetchone()
            id_cliente = fila[0] if fila else None

        if not id_cliente:
            resultados["errores"].append(f"Fila {num} '{nombre}': cliente no encontrado")
            continue

        info_plan = PLANES_FIVGYM.get(plan, {"duracion": 30, "precio": sus["total"]})
        try:
            id_mem = obtener_o_crear_mem_local(plan, info_plan["precio"], info_plan["duracion"])
        except Exception as e:
            resultados["errores"].append(f"Membresia '{plan}': {e}")
            continue

        # FIX: si no hay pagos registrados en el Excel, pagado = 0
        # Antes: si pago1+pago2 = 0, se asignaba total como pagado (incorrecto)
        suma_pagos = sus["pago1"] + sus["pago2"]
        pagado     = suma_pagos  # 0 si no hay pagos, o el monto real si los hay
        pendiente  = max(0, sus["total"] - pagado)

        try:
            con.execute("""
                INSERT OR IGNORE INTO suscripciones
                    (cliente_id, membresia_id, fecha_inicio, fecha_vencimiento,
                     precio_total, pagado, pendiente)
                VALUES (?,?,?,?,?,?,?)
            """, (id_cliente, id_mem, sus["inicio"], sus["fin"],
                  sus["total"], pagado, pendiente))
            resultados["suscripciones"] += 1
        except Exception as e:
            resultados["errores"].append(f"Suscripcion fila {num} '{nombre}': {e}")

    con.execute("PRAGMA foreign_keys = ON")
    con.commit()
    con.close()
    return resultados


# ── DETECTAR FORMATO ──────────────────────────────────────────────────────────

def detectar_formato(ruta: str) -> str:
    wb = openpyxl.load_workbook(ruta, data_only=True)
    hojas = wb.sheetnames
    if "DATOS INFORMATIVOS" in hojas and "PLANES Y PAGOS" in hojas:
        return "fivgym"
    if "Clientes" in hojas or "Membresias" in hojas:
        return "plantilla"
    return "desconocido"


# ── VENTANA DE IMPORTACIÓN ────────────────────────────────────────────────────

def abrir_ventana_importar(parent):
    popup = ttk.Toplevel()
    popup.title("Importar Excel")
    popup.geometry("500x360")
    popup.resizable(False, False)

    parent.update_idletasks()
    x = parent.winfo_x() + (parent.winfo_width()  // 2) - 250
    y = parent.winfo_y() + (parent.winfo_height() // 2) - 180
    popup.geometry(f"+{x}+{y}")
    popup.lift()

    # ── Cabecera fija ────────────────────────────────────────────────────────
    frame_header = ttk.Frame(popup, padding=(20, 16, 20, 0))
    frame_header.pack(fill="x")

    ttk.Label(frame_header, text="Importar Excel",
              font=("Segoe UI", 16, "bold")).pack()
    ttk.Label(frame_header,
              text="Selecciona un archivo Excel. El formato se detecta automaticamente.",
              font=("Segoe UI", 10), justify="center").pack(pady=(4, 10))
    ttk.Separator(frame_header).pack(fill="x")

    # ── Zona con scrollbar ───────────────────────────────────────────────────
    scroll_frame = ttk.Frame(popup)
    scroll_frame.pack(fill="both", expand=True)

    sb = ttk.Scrollbar(scroll_frame, orient="vertical")
    sb.pack(side="right", fill="y")

    canvas_pop = ttk.Canvas(scroll_frame, highlightthickness=0, yscrollcommand=sb.set)
    canvas_pop.pack(side="left", fill="both", expand=True)
    sb.configure(command=canvas_pop.yview)

    inner = ttk.Frame(canvas_pop, padding=(20, 14, 20, 14))
    win_id = canvas_pop.create_window((0, 0), window=inner, anchor="nw")

    def _ajustar(event):
        canvas_pop.configure(scrollregion=canvas_pop.bbox("all"))
        canvas_pop.itemconfig(win_id, width=canvas_pop.winfo_width())
    inner.bind("<Configure>", _ajustar)

    ttk.Label(inner, text="Formatos soportados:",
              font=("Segoe UI", 11, "bold")).pack(anchor="w", pady=(0, 6))
    ttk.Label(inner, text="  Plantilla estandar  (plantilla_gym.xlsx)",
              font=("Segoe UI", 10)).pack(anchor="w", pady=2)
    ttk.Label(inner, text="  Formato FIVGYM  ",
              font=("Segoe UI", 10)).pack(anchor="w")

    ttk.Separator(inner).pack(fill="x", pady=10)

    lbl_estado = ttk.Label(inner, text="", font=("Segoe UI", 10), justify="center")
    lbl_estado.pack(pady=(0, 10))

    def seleccionar_e_importar():
        ruta = filedialog.askopenfilename(
            parent=popup,
            title="Selecciona el archivo Excel",
            filetypes=[("Excel", "*.xlsx *.xls"), ("Todos", "*.*")]
        )
        if not ruta:
            return

        lbl_estado.configure(text="Procesando...")
        popup.update()

        try:
            formato = detectar_formato(ruta)

            if formato == "plantilla":
                res = importar_plantilla(ruta)
                msg = (
                    f"Importacion completada:\n\n"
                    f"  Clientes importados:   {res['clientes']}\n"
                    f"  Membresias importadas: {res['membresias']}\n"
                )
            elif formato == "fivgym":
                res = importar_fivgym(ruta)
                msg = (
                    f"Importacion completada:\n\n"
                    f"  Clientes importados:      {res['clientes']}\n"
                    f"  Fichas importadas:        {res['fichas']}\n"
                    f"  Suscripciones importadas: {res['suscripciones']}\n"
                )
            else:
                lbl_estado.configure(text="Formato no reconocido.")
                messagebox.showerror(
                    "Formato no reconocido",
                    "El archivo no es compatible.\nUsa la plantilla estandar o TABLA_FIVGYM.",
                    parent=popup
                )
                return

            errores = res.get("errores", [])
            if errores:
                msg += f"\nAdvertencias ({len(errores)}):\n"
                for e in errores[:5]:
                    msg += f"  - {e}\n"
                if len(errores) > 5:
                    msg += f"  ... y {len(errores) - 5} mas.\n"

            lbl_estado.configure(text="Importacion exitosa")
            messagebox.showinfo("Listo", msg, parent=popup)

        except Exception as e:
            lbl_estado.configure(text=f"Error: {e}")
            messagebox.showerror("Error", f"No se pudo importar:\n{e}", parent=popup)

    ttk.Button(
        inner,
        text="Seleccionar archivo Excel",
        bootstyle="primary",
        width=28,
        padding=10,
        command=seleccionar_e_importar
    ).pack(pady=(0, 8))

    # ── Footer fijo ──────────────────────────────────────────────────────────
    frame_footer = ttk.Frame(popup, padding=(20, 8, 20, 14))
    frame_footer.pack(fill="x")
    ttk.Separator(frame_footer).pack(fill="x", pady=(0, 10))
    ttk.Button(frame_footer, text="Cerrar", bootstyle="secondary-outline",
               width=12, command=popup.destroy).pack()